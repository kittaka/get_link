#!/usr/bin/env python
# coding: utf-8

# In[3]:


import time
from msedge.selenium_tools import Edge, EdgeOptions
from selenium import webdriver
from bs4 import BeautifulSoup
import openpyxl
import pprint

def get_link(link,condition):
    # Launch Microsoft Edge (Chromium)
    options = EdgeOptions()
    options.use_chromium = True
    options.binary_location = r"C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedge.exe"
    #options.add_argument("headless")        #Chrome と合わせて欲しいぞ・・・
    #options.add_argument("disable-gpu")
    driver = Edge(options = options)
    
    driver.get(link)

    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    title_text = soup.find('title').get_text()
    print(title_text)

    links = [url.get('href') for url in soup.select(condition)]
    
    driver.quit()
    return links

def get_item(link,condition):
    retS=[]
    print(link)
    options = EdgeOptions()
    options.use_chromium = True
    options.binary_location = r"C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedge.exe"
    driver = Edge(options = options)

    driver.get(link)
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')

    items = [n.get_text() for n in soup.select(condition)]
    
    driver.quit()

    return items
        
def excel_items_write(items,cnty,cntx):
    cntx=int(cntx)
    cnty=int(cnty)
    wb = openpyxl.load_workbook('getdata.xlsx')
    ws = wb.worksheets[0]
    file_name = "getdata.xlsx"
        
    for item in items:
        c1 = ws.cell(cnty, cntx)
        c1.value= item
        cntx+=1
    wb.save(file_name)
            
def excel_read(cnty,cntx):
    wb = openpyxl.load_workbook('getdata.xlsx')
    ws = wb.worksheets[0]
    cnty=int(cnty)
    cntx=int(cntx)
    ret=[]
    i=0
    
    while ws.cell(cnty+i, cntx).value is not None:
        #print(ws.cell(cnty+i, cntx).value)
        ret.append(ws.cell(cnty+i, cntx).value)
        i+=1
    return ret

def excel_links_write(links,cnty,cntx):
    cntx=int(cntx)
    cnty=int(cnty)
    wb = openpyxl.load_workbook('getdata.xlsx')
    ws = wb.worksheets[0]
    file_name = "getdata.xlsx"
        
    for link in links:
        c1 = ws.cell(cnty, cntx)
        c1.value= link
        cnty+=1
    wb.save(file_name)

if __name__ == "__main__":
    # テスト
    print("メイン実行")
    
    


# In[ ]:




